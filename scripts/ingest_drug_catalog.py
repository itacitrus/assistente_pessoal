#!/usr/bin/env python3
"""Ingestao do catalogo de medicamentos (Lista CMED/ANVISA) para drug_catalog.

Fonte: Lista de Precos de Medicamentos (PMC) da CMED/ANVISA, em XLSX. E a unica
fonte publica que traz a APRESENTACAO (concentracao/forma) junto do nome
comercial e do principio ativo — o catalogo de registros (DADOS_ABERTOS) nao tem
dose. Cada linha do XLSX e uma apresentacao especifica (ex: "Losartana 50 MG" e
"Losartana 100 MG" sao linhas distintas), o que alimenta direto o autocomplete
do cadastro de remedio.

O bot (Go) apenas LE drug_catalog; a resolucao fuzzy/fonetica vive em
bot/drug_catalog.go. Este script e o unico que ESCREVE — roda fora do runtime,
quando se quer atualizar o catalogo (a CMED publica uma edicao nova por mes).

Idempotente: recria o conteudo inteiro numa transacao (DELETE + reinsert), entao
pode ser reexecutado a vontade.

Uso:
    python3 scripts/ingest_drug_catalog.py                 # baixa a CMED e popula bot/data/bot.db
    python3 scripts/ingest_drug_catalog.py --xlsx /tmp/cmed.xlsx   # usa arquivo local
    python3 scripts/ingest_drug_catalog.py --db /caminho/bot.db --url <URL CMED>

Dependencia: openpyxl (pip install openpyxl).
"""

from __future__ import annotations

import argparse
import os
import re
import ssl
import sqlite3
import sys
import unicodedata
import urllib.request
from datetime import datetime, timezone

# URL da edicao vigente da Lista CMED (PMC). Muda a cada mes — o nome do arquivo
# carrega a data de publicacao. Atualizar quando rodar uma nova ingestao, ou
# passar --url. Pagina oficial:
# https://www.gov.br/anvisa/pt-br/assuntos/medicamentos/cmed/precos
DEFAULT_CMED_URL = (
    "https://www.gov.br/anvisa/pt-br/assuntos/medicamentos/cmed/precos/"
    "arquivos/xls_conformidade_site_20260508_234642408.xlsx/@@download/file"
)

# Indices de coluna (0-based) na planilha CMED, a partir da linha de cabecalho
# (localizada dinamicamente por SUBSTANCIA). Estaveis ha varias edicoes.
COL_SUBSTANCIA = 0
COL_LABORATORIO = 2
COL_REGISTRO = 4
COL_EAN1 = 5
COL_PRODUTO = 8
COL_APRESENTACAO = 9
COL_CLASSE = 10
COL_TIPO = 11
COL_TARJA = 72

# Unidades de dose conhecidas (uppercase, sem espaco). Usadas para reconhecer o
# prefixo de concentracao no inicio da APRESENTACAO. Lista nao exaustiva: se uma
# unidade exotica nao estiver aqui, a concentracao apenas para mais cedo — a
# apresentacao crua continua preservada na coluna `presentation`.
UNIT_WORDS = {
    "MG", "G", "MCG", "UI", "ML", "L", "KG", "%", "MEQ", "MMOL", "MUI",
    "MG/ML", "MG/G", "MCG/ML", "MCG/G", "MG/DOSE", "MCG/DOSE", "UI/ML",
    "G/ML", "MG/L", "UI/G", "MG/MG", "MG/ML/ML", "MCG/JATO", "MG/JATO",
}


def normalize(text: str) -> str:
    """minusculo, sem acento, sem pontuacao, espacos colapsados."""
    if not text:
        return ""
    nfkd = unicodedata.normalize("NFKD", text)
    no_accent = "".join(c for c in nfkd if not unicodedata.combining(c))
    lowered = no_accent.lower()
    cleaned = re.sub(r"[^a-z0-9 ]+", " ", lowered)
    return re.sub(r"\s+", " ", cleaned).strip()


_NUM = re.compile(r"^[\d.,]+$")
_NUM_UNIT = re.compile(r"^[\d.,]+[A-Z%][A-Z%/]*$")


def is_dose_token(tok: str) -> bool:
    t = tok.upper()
    if t == "+":
        return True
    if _NUM.match(t):
        return True
    if t in UNIT_WORDS:
        return True
    if _NUM_UNIT.match(t):
        return True
    return False


def extract_concentration(apresentacao: str) -> str:
    """Extrai o prefixo de dose da APRESENTACAO ("50 MG COM REV..." -> "50 MG").

    Consome tokens iniciais enquanto parecem dose (numero, unidade, "+",
    numero+unidade fundidos) e para no primeiro token de forma/embalagem.
    """
    if not apresentacao:
        return ""
    tokens = apresentacao.strip().split()
    taken: list[str] = []
    for tok in tokens:
        if is_dose_token(tok):
            taken.append(tok)
        else:
            break
    # Nao deixar um "+" pendurado no fim ("10 MG +" -> "10 MG").
    while taken and taken[-1] == "+":
        taken.pop()
    return " ".join(taken)


def find_header_row(rows) -> int:
    """Localiza o indice (0-based) da linha de cabecalho pela celula SUBSTANCIA.

    A CMED traz ~40 linhas de legenda antes do cabecalho, e o numero varia entre
    edicoes — por isso buscamos em vez de fixar.
    """
    for i, row in enumerate(rows):
        first = row[0]
        if first is not None and normalize(str(first)) == "substancia":
            return i
        if i > 200:  # sanity guard
            break
    raise RuntimeError("cabecalho da CMED nao encontrado (coluna SUBSTANCIA)")


# DDL espelhado de bot/db.go (migrate). CREATE IF NOT EXISTS para o script ser
# auto-suficiente num banco recem-criado. Manter em sincronia com o Go.
CREATE_TABLE = """
CREATE TABLE IF NOT EXISTS drug_catalog (
    id                INTEGER PRIMARY KEY AUTOINCREMENT,
    commercial_name   TEXT NOT NULL,
    active_ingredient TEXT NOT NULL DEFAULT '',
    concentration     TEXT NOT NULL DEFAULT '',
    presentation      TEXT NOT NULL DEFAULT '',
    lab               TEXT NOT NULL DEFAULT '',
    ean               TEXT NOT NULL DEFAULT '',
    anvisa_reg        TEXT NOT NULL DEFAULT '',
    product_type      TEXT NOT NULL DEFAULT '',
    tarja             TEXT NOT NULL DEFAULT '',
    therapeutic_class TEXT NOT NULL DEFAULT '',
    norm_name         TEXT NOT NULL DEFAULT '',
    norm_ingredient   TEXT NOT NULL DEFAULT '',
    source_version    TEXT NOT NULL DEFAULT '',
    ingest_token      TEXT NOT NULL DEFAULT '',
    created_at        DATETIME DEFAULT CURRENT_TIMESTAMP
)
"""

# Chave natural estavel para o upsert. PRECISA bater com o indice UNIQUE criado
# pelo Go (idx_drug_catalog_natkey). Mantem os ids entre edicoes da CMED, para
# que medications.catalog_id continue valido apos uma reingestao.
CREATE_NATKEY_INDEX = (
    "CREATE UNIQUE INDEX IF NOT EXISTS idx_drug_catalog_natkey "
    "ON drug_catalog(norm_name, concentration)"
)


def cell(row, idx: int) -> str:
    if idx >= len(row):
        return ""
    v = row[idx]
    if v is None:
        return ""
    return str(v).strip()


def clean_tarja(raw: str) -> str:
    """Normaliza a TARJA da CMED para "" | "Vermelha" | "Preta".

    A CMED grava "Tarja Vermelha", "Tarja Preta", "Tarja Sem Tarja" e "- (*)".
    Tiramos o prefixo "Tarja " e tratamos "sem tarja"/"-"/vazio como sem tarja
    (medicamento de venda livre). A tarja sinaliza controle (receita/retencao).
    """
    t = raw.strip()
    if not t or t.startswith("-"):
        return ""
    t = re.sub(r"(?i)^tarja\s+", "", t).strip()
    if normalize(t) in ("sem tarja", ""):
        return ""
    return t


def download(url: str, dest: str) -> None:
    # O servidor da ANVISA serve a cadeia de certificados de forma incompleta;
    # urllib falha na verificacao. Mantemos a verificacao desligada apenas para
    # este download de dado publico (sem credenciais, sem PII).
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    print(f"baixando CMED de {url} ...", file=sys.stderr)
    req = urllib.request.Request(url, headers={"User-Agent": "zello-ingest/1.0"})
    with urllib.request.urlopen(req, context=ctx, timeout=300) as resp, open(dest, "wb") as fh:
        fh.write(resp.read())
    print(f"salvo em {dest} ({os.path.getsize(dest)/1e6:.1f} MB)", file=sys.stderr)


def main() -> int:
    repo_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    default_db = os.path.join(repo_root, "bot", "data", "bot.db")

    ap = argparse.ArgumentParser(description="Ingestao do catalogo CMED em drug_catalog")
    ap.add_argument("--db", default=default_db, help=f"caminho do SQLite (default: {default_db})")
    ap.add_argument("--xlsx", default=None, help="arquivo XLSX local (pula o download)")
    ap.add_argument("--url", default=DEFAULT_CMED_URL, help="URL da Lista CMED (PMC) XLSX")
    args = ap.parse_args()

    try:
        import openpyxl  # noqa: WPS433 (import tardio: dependencia so do ingest)
    except ImportError:
        print("ERRO: openpyxl nao instalado. Rode: pip install openpyxl", file=sys.stderr)
        return 2

    xlsx_path = args.xlsx
    tmp_download = None
    if not xlsx_path:
        tmp_download = os.path.join("/tmp", "cmed_ingest.xlsx")
        download(args.url, tmp_download)
        xlsx_path = tmp_download

    print(f"lendo {xlsx_path} ...", file=sys.stderr)
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)

    # iter_rows e um gerador de uso unico; materializamos para achar o header e
    # depois iterar os dados. ~25k linhas x poucas colunas relevantes cabe bem.
    all_rows = list(rows)
    header_idx = find_header_row(all_rows)
    source_version = ""
    for r in all_rows[:header_idx]:
        if r and r[0] and "publicada" in normalize(str(r[0])):
            source_version = str(r[0]).strip()
            break

    # Token unico deste run: marca toda linha inserida/atualizada agora. No fim,
    # linhas com token antigo (apresentacoes que sairam da edicao da CMED) sao
    # removidas. Garante limpeza sem precisar de DELETE-tudo (que reatribuiria ids).
    run_token = f"{datetime.now(timezone.utc).isoformat()}/{os.urandom(4).hex()}"

    seen: set[tuple[str, str]] = set()
    records: list[tuple] = []
    skipped_no_name = 0
    for row in all_rows[header_idx + 1:]:
        produto = cell(row, COL_PRODUTO)
        if not produto:
            skipped_no_name += 1
            continue
        apresentacao = cell(row, COL_APRESENTACAO)
        concentration = extract_concentration(apresentacao)
        norm_name = normalize(produto)
        # Dedup: colapsa variantes de laboratorio/embalagem da mesma apresentacao.
        # Para o lembrete so importa nome + concentracao; mantemos o 1o registro
        # como representante (com seu EAN/lab/registro). A concentracao entra na
        # chave SEM espacos para que "500 MG" e "500MG" (grafias divergentes na
        # origem) colapsem na mesma entrada.
        dedup_key = (norm_name, normalize(concentration).replace(" ", ""))
        if dedup_key in seen:
            continue
        seen.add(dedup_key)

        substancia = cell(row, COL_SUBSTANCIA)
        records.append((
            produto,
            substancia,
            concentration,
            apresentacao,
            cell(row, COL_LABORATORIO),
            cell(row, COL_EAN1),
            cell(row, COL_REGISTRO),
            cell(row, COL_TIPO),
            clean_tarja(cell(row, COL_TARJA)),
            cell(row, COL_CLASSE),
            norm_name,
            normalize(substancia),
            source_version,
            run_token,
        ))
    wb.close()

    print(
        f"parse: {len(records)} apresentacoes unicas "
        f"(dedup colapsou variantes; {skipped_no_name} linhas sem nome ignoradas)",
        file=sys.stderr,
    )

    os.makedirs(os.path.dirname(os.path.abspath(args.db)), exist_ok=True)
    conn = sqlite3.connect(args.db)
    try:
        conn.execute(CREATE_TABLE)
        conn.execute(CREATE_NATKEY_INDEX)
        with conn:  # transacao: tudo-ou-nada
            # Upsert por chave natural (norm_name, concentration): novas entradas
            # ganham id; entradas existentes mantem o id e tem os campos
            # atualizados. created_at NAO eh sobrescrito (preserva quando a
            # entrada apareceu pela 1a vez).
            conn.executemany(
                """INSERT INTO drug_catalog (
                    commercial_name, active_ingredient, concentration, presentation,
                    lab, ean, anvisa_reg, product_type, tarja, therapeutic_class,
                    norm_name, norm_ingredient, source_version, ingest_token
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(norm_name, concentration) DO UPDATE SET
                    commercial_name   = excluded.commercial_name,
                    active_ingredient = excluded.active_ingredient,
                    presentation      = excluded.presentation,
                    lab               = excluded.lab,
                    ean               = excluded.ean,
                    anvisa_reg        = excluded.anvisa_reg,
                    product_type      = excluded.product_type,
                    tarja             = excluded.tarja,
                    therapeutic_class = excluded.therapeutic_class,
                    norm_ingredient   = excluded.norm_ingredient,
                    source_version    = excluded.source_version,
                    ingest_token      = excluded.ingest_token""",
                records,
            )
            # Sweep: remove apresentacoes que nao vieram nesta edicao.
            cur = conn.execute(
                "DELETE FROM drug_catalog WHERE ingest_token != ?", (run_token,)
            )
            removed = cur.rowcount
        total = conn.execute("SELECT COUNT(*) FROM drug_catalog").fetchone()[0]
        print(
            f"OK: {total} medicamentos em drug_catalog ({args.db}) | "
            f"upsert {len(records)} | removidas (fora desta edicao) {removed} | "
            f"fonte: {source_version or 'CMED'} | ingerido {datetime.now(timezone.utc).isoformat()}",
            file=sys.stderr,
        )
    finally:
        conn.close()

    if tmp_download and os.path.exists(tmp_download):
        os.remove(tmp_download)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
